#생년 나이 계산 시트 만들기

import openpyxl as excel
import datetime

book = excel.Workbook()
sheet = book.active

thisyear = datetime.datetime.now().year
#print(thisyear)

sheet["A1"] = "Birth year"
sheet["B1"] = "Korea age"
sheet["C1"] = "Global age(before birthday)"
sheet["D1"] = "Global age(after birthday)"
sheet["E1"] = "to 100 years old(kr_age)"
sheet["F1"] = "to go elementary school"



sheet.column_dimensions["C"].width=25
sheet.column_dimensions["D"].width=25
sheet.column_dimensions["E"].width=25
sheet.column_dimensions["F"].width=25

for i in range(81):
    birth_year = thisyear - i
    korea_age = thisyear - birth_year + 1
    global_age ={"after_bday":korea_age -1 , "before_bday":korea_age -2}
    hundred_age = 100 - korea_age
    school_age = birth_year + 8

    year_cell=sheet.cell(i+2, 1)
    year_cell.value = str(birth_year) + "년생"

    age_cell = sheet.cell(i+2, 2)
    age_cell.value = str(korea_age) + "세"

    age_cell = sheet.cell(i+2, 3)
    age_cell.value = "만" + str(global_age["after_bday"]) + "세"

    age_cell = sheet.cell(i+2, 4)
    age_cell.value = "만" + str(global_age["before_bday"]) + "세"

    age_cell = sheet.cell(i+2, 5)
    age_cell.value = str(hundred_age) + "년"

    age_cell = sheet.cell(i+2, 6)
    age_cell.value = str(school_age) + "년" 

sheet["D2"] = "-"

book.save("./output/write_agelist.xlsx")



