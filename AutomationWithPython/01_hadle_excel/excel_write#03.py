#구구단 입력하기

from pyexpat.model import XML_CTYPE_CHOICE
import openpyxl as excel

book = excel.Workbook()
sheet = book.active

for y in range(1,10):
    for x in range(1,10):
        #cell 읽기
        cell = sheet.cell(row=y, column=x)
        #value 입력
        cell.value = x * y

book.save("./output/write_9x9.xlsx")



for a in range(1,100):
    for b in range(1,100):
        cell = sheet.cell(a,b)
        cell.value = a*b
        
book.save("./output/write_99x99.xlsx")
