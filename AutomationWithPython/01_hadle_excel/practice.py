# import openpyxl as excel

# book = excel.Workbook()
# sheet = book.active
# cell = sheet.cell(row=1,column=1)
# cell.value = "abc"

# print(cell.value)

# cell2 = sheet.cell(row=1,column=2,value="def")

# print (cell2.value)

# cell3 = sheet.cell(row=1,column=3)
# cell3.value = cell3.coordinate
# print(cell3.value)

import openpyxl as excel
book = excel.load_workbook("./output/write_cellname.xlsx")
sheet = book.active

# print(sheet["H12"].value)
# cell = sheet.cell(row=3, column=13)
# print (cell.value)

for y in range (2, 5):
    r = []
    for x in range (2, 5):
        v = sheet.cell(row=y, column=x).value
        # row=B , column = 2
        # row=B , column = 3
        # row=B , column = 4
        r.append(v)
    print(r)