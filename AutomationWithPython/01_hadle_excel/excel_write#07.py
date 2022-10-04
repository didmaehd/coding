#연도별 육갑십자 구하기 

import openpyxl as excel
import datetime

#간지와 지지 리스트로 저장
gan = ["갑을병정무기경신임계","청청적적황황백백흑흑"]
ji = ["자축인묘진사오미신유술해",["쥐","소","호랑이","토끼","용","뱀","말","양","원숭이","닭","개","돼지"]]

#연도별 간지 지지 구하기
def year_to_ganji(year):
    i = (year-4)%10 # 서기 4년이 갑자년이므로 입력 연도에서 4를 뺀후 10으로 나눈 나머지로 간지를 구할수 있다,
    j = (year-4)%12 # 서기 4년이 갑자년이므로 입력 연도에서 4를 뺀후 12로 나는 나머지로 지지를 구할수 있다,

    ganji = gan[0][i] + ji[0][j] +"년" #나머지 값으로 리스트에서 간지, 지지 뽑아서 ganji변수에 넣는다
    color_animal = gan[1][i] +"색"+ ji[1][j] # 나머지 값으로 리스트에서 동물의 색을 뽑아서 변수에 넣는다.
    return ganji, color_animal  # 간지와 동물색을 리턴한다

abc = year_to_ganji(1592)
print (abc)

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "Year"
sheet["B1"] = "Gnaji"
sheet["C1"] = "Animal"

# start_y = now = datetime.datetime.now().year
start_y = 2050

for i in range(100):
    #2050년 기준으로 과거 100년간 간지와 동물, 색 구하기
    year = start_y -i #2050년 기준으로 과거 100년까지 반복
    result = year_to_ganji(year) # 연도값 100년치를 위 함수에 돌려서 간지와 동물색을 불러온다
    ganji = result[0] # 결과값의 첫번째 값 간지를 변수에 저장
    col_ani = result[1] # 결과값의 두번째 값 동물색을 변수에 저장
    #결과값을 엑셀시트에 입력
    sheet.cell(i+2 , 1, value = str(year) + "년")
    sheet.cell(i+2 , 2, value = ganji)
    sheet.cell(i+2 , 3, value = col_ani)
    print (year, "=", ganji , ",", col_ani)

#엑셀 파일 저장
book.save("./output/write_ganji.xlsx")