#셀에 값을 넣는 세가지 방법

from colorsys import TWO_THIRD
import openpyxl as excel
#워크북 생성
book = excel.Workbook()
sheet = book.active

#1 셀 주소를 지정해서 값을 넣는 방법
sheet["A1"] = "Today, I studied AWS and python for a whole day."

#2 row , column, value 를 지정해 값을 입력
sheet.cell(row=2, column = 1, value="So, I am very perfectly tired.")

#3 row와 column을 변수에 넣은후 value를 추가 입력
third_cell = sheet.cell(row=3, column=1)
third_cell.value = "But I thought, learning something is certainly a worthwhile thing"

#파일 저장
book.save("./output/diary.xlsx")