import openpyxl as excel
import datetime

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "birth year"
sheet["B1"] = "Enter Elementary School"
sheet["C1"] = "University ID"

sheet.column_dimensions["A"].width = 50
sheet.column_dimensions["B"].width = 50
sheet.column_dimensions["C"].width = 50

for i in range(50):
    #birth year
    birth_year = 2002 - i
    birth_range = (f"{birth_year} Mar 1 ~ {birth_year+1} Feb 28(29)")

    #Start year a Elementary school 
    ele_school = birth_year + 7

    #University ID number
    univ_year = birth_year + 19
    univ_num = str(univ_year)[2:]

    #insert data
    sheet.cell(i+2,1,value=birth_range)
    sheet.cell(i+2,2,value=str(ele_school)+" year")
    sheet.cell(i+2,3,value=univ_num)

#exception handling
sheet["A2"] = "2002 Mar 1 ~ 2002 Dec 31"

book.save("./output/write_entry_year.xlsx")

